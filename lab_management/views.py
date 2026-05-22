from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django import forms
from django.conf import settings
from django.http import FileResponse, HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import models
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
import os
import json
import hmac
import hashlib
import base64
import time
from datetime import datetime
import requests
from .models import UserProfile, Document, DailyPlan, Announcement, AnnouncementComment, Achievement, Message, ResumeDraft


class UserRegistrationForm(forms.ModelForm):
    password = forms.CharField(widget=forms.PasswordInput)
    password_confirm = forms.CharField(widget=forms.PasswordInput)
    
    class Meta:
        model = User
        fields = ['username', 'email', 'first_name']
    
    def clean_password_confirm(self):
        password = self.cleaned_data.get('password')
        password_confirm = self.cleaned_data.get('password_confirm')
        if password and password_confirm and password != password_confirm:
            raise forms.ValidationError("两次输入的密码不一致")
        return password_confirm


def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return render(request, 'lab_management/register.html', {'form': form})
        
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            is_staff=True
        )
        
        UserProfile.objects.create(
            user=user,
            research_topic='待填写',
            bio=''
        )
        
        user = authenticate(username=username, password=password)
        login(request, user)
        messages.success(request, '注册成功！请完善您的个人资料')
        return redirect('my_page')
    else:
        form = UserRegistrationForm()
    
    return render(request, 'lab_management/register.html', {'form': form})


def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'home')
            return redirect(next_url)
        else:
            messages.error(request, '用户名或密码错误')
    
    return render(request, 'lab_management/login.html')


def user_logout(request):
    logout(request)
    return redirect('home')


def home(request):
    # 优化：使用预加载减少数据库查询
    members = UserProfile.objects.select_related('user').filter(user__is_superuser=False).all()
    
    # 获取所有计划（包括非公开的），预加载子计划
    all_plans = DailyPlan.objects.select_related('user').prefetch_related('sub_plans').filter(
        user__is_superuser=False, parent__isnull=True
    )
    
    # 获取所有公开计划
    public_plans = all_plans.filter(is_public=True)
    
    announcements = Announcement.objects.all()[:3]
    announcement_count = Announcement.objects.count()
    public_documents = Document.objects.select_related('user').filter(is_public=True)
    document_count = public_documents.count()
    
    # 获取当前用户的所有计划
    if request.user.is_authenticated:
        if request.user.is_superuser:
            user_own_plans = all_plans
        else:
            user_own_plans = all_plans.filter(user=request.user)
    else:
        user_own_plans = []
    
    # 优化：一次性获取所有计划数据，避免循环内查询
    # 构建用户ID到计划的映射
    user_plans_map = {}
    for plan in all_plans:
        user_id = plan.user_id
        if user_id not in user_plans_map:
            user_plans_map[user_id] = []
        user_plans_map[user_id].append(plan)
    
    # 构建公开计划映射
    public_plans_map = {}
    for plan in public_plans:
        user_id = plan.user_id
        if user_id not in public_plans_map:
            public_plans_map[user_id] = []
        public_plans_map[user_id].append(plan)
    
    # 为每个成员构建计划数据（使用预加载的数据）
    members_with_plans = []
    for member in members:
        user_id = member.user_id
        
        # 根据用户权限选择计划
        if request.user.is_authenticated and (user_id == request.user.id or request.user.is_superuser):
            user_plans_all = user_plans_map.get(user_id, [])
        else:
            user_plans_all = public_plans_map.get(user_id, [])
        
        user_plans = user_plans_all[:3]
        
        # 使用预加载的子计划数据（避免额外查询）
        for plan in user_plans:
            if request.user.is_authenticated and (plan.user_id == request.user.id or request.user.is_superuser):
                plan.public_sub_plans = plan.sub_plans.all()
            else:
                plan.public_sub_plans = [sp for sp in plan.sub_plans.all() if sp.is_public]
            plan.public_sub_plans_count = len(plan.public_sub_plans)
        
        members_with_plans.append({
            'profile': member,
            'plans': user_plans,
            'plan_count': len(user_plans_all)
        })
    
    # 统计近期公开计划总数
    recent_plan_count = len(list(public_plans))
    
    context = {
        'members': members,
        'members_with_plans': members_with_plans,
        'recent_plan_count': recent_plan_count,
        'announcements': announcements,
        'announcement_count': announcement_count,
        'document_count': document_count,
        'recent_documents': public_documents[:5],
    }
    return render(request, 'lab_management/home.html', context)


def member_list(request):
    member_type_filter = request.GET.get('member_type', 'all')
    query = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'grade')
    view = request.GET.get('view', 'grid')
    valid_types = {'all', 'phd', 'master', 'undergraduate', 'graduated'}
    valid_sorts = {'grade', 'plans', 'updated'}
    valid_views = {'grid', 'list'}
    if member_type_filter not in valid_types:
        member_type_filter = 'all'
    if sort not in valid_sorts:
        sort = 'grade'
    if view not in valid_views:
        view = 'grid'

    # 获取所有非管理员成员，统计数据基于全量，列表基于筛选结果
    all_members = UserProfile.objects.select_related('user').filter(user__is_superuser=False)

    active_plan_counts = {}
    completed_plan_counts = {}
    latest_plan_dates = {}
    for row in DailyPlan.objects.filter(
        user__is_superuser=False,
        parent__isnull=True,
        is_public=True,
        is_completed=False,
    ).values('user_id').annotate(count=models.Count('id')):
        active_plan_counts[row['user_id']] = row['count']
    for row in DailyPlan.objects.filter(
        user__is_superuser=False,
        parent__isnull=True,
        is_public=True,
        is_completed=True,
    ).values('user_id').annotate(count=models.Count('id')):
        completed_plan_counts[row['user_id']] = row['count']
    for row in DailyPlan.objects.filter(
        user__is_superuser=False,
        parent__isnull=True,
        is_public=True,
    ).values('user_id').annotate(latest=models.Max('date')):
        latest_plan_dates[row['user_id']] = row['latest']

    phd_members = all_members.filter(member_type='phd')
    master_members = all_members.filter(member_type='master')
    undergraduate_members = all_members.filter(member_type='undergraduate')
    graduated_members = all_members.filter(member_type='graduated')
    master_grade1 = master_members.filter(grade='1')
    master_grade2 = master_members.filter(grade='2')
    master_grade3 = master_members.filter(grade='3')

    filtered_members = all_members
    if member_type_filter != 'all':
        filtered_members = filtered_members.filter(member_type=member_type_filter)
    if query:
        filtered_members = filtered_members.filter(
            models.Q(user__first_name__icontains=query) |
            models.Q(user__username__icontains=query) |
            models.Q(research_topic__icontains=query) |
            models.Q(bio__icontains=query) |
            models.Q(email__icontains=query)
        )

    member_cards = []
    for profile in filtered_members:
        active_count = active_plan_counts.get(profile.user_id, 0)
        completed_count = completed_plan_counts.get(profile.user_id, 0)
        total_count = active_count + completed_count
        member_cards.append({
            'profile': profile,
            'project_count': active_count,
            'completed_count': completed_count,
            'total_count': total_count,
            'completion_rate': round(completed_count / total_count * 100) if total_count else 0,
            'latest_update': latest_plan_dates.get(profile.user_id) or profile.user.date_joined.date(),
        })

    type_order = {'phd': 0, 'master': 1, 'undergraduate': 2, 'graduated': 3}
    if sort == 'plans':
        member_cards.sort(key=lambda item: (item['project_count'], item['completed_count'], item['profile'].user_id), reverse=True)
    elif sort == 'updated':
        member_cards.sort(key=lambda item: (item['latest_update'], item['profile'].user_id), reverse=True)
    else:
        member_cards.sort(key=lambda item: (
            type_order.get(item['profile'].member_type, 9),
            int(item['profile'].grade) if str(item['profile'].grade).isdigit() else 9,
            item['profile'].user.first_name or item['profile'].user.username
        ))
    
    context = {
        'all_members': all_members,
        'member_cards': member_cards,
        'phd_members': phd_members,
        'master_members': master_members,
        'undergraduate_members': undergraduate_members,
        'graduated_members': graduated_members,
        'master_grade1': master_grade1,
        'master_grade2': master_grade2,
        'master_grade3': master_grade3,
        'active_project_total': sum(active_plan_counts.values()),
        'member_type_filter': member_type_filter,
        'query': query,
        'sort': sort,
        'view': view,
    }
    return render(request, 'lab_management/member_list.html', context)


def plan_list(request):
    from datetime import datetime, timedelta
    import calendar as cal
    
    view = request.GET.get('view', 'list')
    if view not in {'list', 'calendar', 'board'}:
        view = 'list'
    requested_member = request.GET.get('member', '')
    selected_user = requested_member
    member_query = request.GET.get('member_q', '').strip()
    plan_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    if status_filter not in {'all', 'in_progress', 'completed', 'delayed'}:
        status_filter = 'all'
    today_date = timezone.localdate()
    
    # 显示所有人公开的计划
    main_plans = DailyPlan.objects.select_related('user', 'user__profile').filter(parent__isnull=True, is_public=True).prefetch_related('sub_plans')
    all_plans = DailyPlan.objects.select_related('user', 'user__profile').filter(is_public=True)
    if requested_member:
        main_plans = main_plans.filter(user__username=selected_user)
        all_plans = all_plans.filter(user__username=selected_user)
    if plan_query:
        main_plans = main_plans.filter(
            models.Q(content__icontains=plan_query) |
            models.Q(user__username__icontains=plan_query) |
            models.Q(user__first_name__icontains=plan_query) |
            models.Q(user__profile__research_topic__icontains=plan_query)
        )
        all_plans = all_plans.filter(
            models.Q(content__icontains=plan_query) |
            models.Q(user__username__icontains=plan_query) |
            models.Q(user__first_name__icontains=plan_query) |
            models.Q(user__profile__research_topic__icontains=plan_query)
        )
    if status_filter == 'completed':
        main_plans = main_plans.filter(is_completed=True)
        all_plans = all_plans.filter(is_completed=True)
    elif status_filter == 'in_progress':
        main_plans = main_plans.filter(is_completed=False, date__gte=today_date)
        all_plans = all_plans.filter(is_completed=False, date__gte=today_date)
    elif status_filter == 'delayed':
        main_plans = main_plans.filter(is_completed=False, date__lt=today_date)
        all_plans = all_plans.filter(is_completed=False, date__lt=today_date)
    
    # 获取所有成员
    members = UserProfile.objects.select_related('user').filter(user__is_superuser=False).all()
    visible_members = members
    if member_query:
        visible_members = visible_members.filter(
            models.Q(user__first_name__icontains=member_query) |
            models.Q(user__username__icontains=member_query) |
            models.Q(research_topic__icontains=member_query)
        )
    all_public_main_plans = DailyPlan.objects.select_related('user').filter(parent__isnull=True, is_public=True, user__is_superuser=False).prefetch_related('sub_plans')
    
    # 为每个成员构建计划数据（包括没有计划的成员）
    members_with_plans = []
    for member in visible_members:
        user_plans = [p for p in main_plans if p.user_id == member.user_id]
        user_all_plans = [p for p in all_public_main_plans if p.user_id == member.user_id]
        
        # 侧栏统计按该成员所有公开主计划计算；右侧内容按当前筛选显示。
        main_completed = len([p for p in user_all_plans if p.is_completed])
        main_total = len(user_all_plans)
        
        total_completed = main_completed
        total_count = main_total
        completion_rate = round(total_completed / total_count * 100, 1) if total_count > 0 else 0
        
        members_with_plans.append({
            'profile': member,
            'plans': user_plans,
            'plan_count': total_count,
            'completed_count': total_completed,
            'in_progress_count': max(total_count - total_completed, 0),
            'completion_rate': completion_rate
        })
    total_plan_count = main_plans.count()
    total_completed_count = main_plans.filter(is_completed=True).count()
    # 延期计划：is_delayed=True 或 deadline < today
    delayed_plan_count = main_plans.filter(
        is_completed=False
    ).filter(
        Q(is_delayed=True) | Q(deadline__isnull=False, deadline__lt=today_date)
    ).count()
    # 进行中计划：未完成且未延期
    total_in_progress_count = main_plans.filter(
        is_completed=False
    ).exclude(
        Q(is_delayed=True) | Q(deadline__isnull=False, deadline__lt=today_date)
    ).count()
    avg_completion_rate = round(total_completed_count / total_plan_count * 100, 0) if total_plan_count > 0 else 0
    selected_member_data = None
    if requested_member:
        selected_member_data = next((item for item in members_with_plans if item['profile'].user.username == selected_user), None)
    if selected_member_data is None and request.user.is_authenticated:
        selected_member_data = next((item for item in members_with_plans if item['profile'].user_id == request.user.id), None)
    if selected_member_data is None:
        selected_member_data = next((item for item in members_with_plans if item['plan_count'] > 0), None)
    if selected_member_data is None and members_with_plans:
        selected_member_data = members_with_plans[0]
    if selected_member_data is not None and not selected_user:
        selected_user = selected_member_data['profile'].user.username

    for plan in main_plans:
        sub_plans = list(plan.sub_plans.filter(is_public=True))
        sub_total = len(sub_plans)
        sub_completed = len([sub for sub in sub_plans if sub.is_completed])
        plan.progress_rate = 100 if plan.is_completed else (round(sub_completed / sub_total * 100) if sub_total else 0)
        plan.sub_completed_count = sub_completed
        plan.visible_sub_plans = sub_plans
        if plan.is_completed:
            plan.ui_status = '已完成'
            plan.ui_status_key = 'completed'
        elif plan.is_delayed or (plan.deadline and plan.deadline < today_date):
            plan.ui_status = '延期'
            plan.ui_status_key = 'delayed'
        else:
            plan.ui_status = '进行中'
            plan.ui_status_key = 'in_progress'

    board_columns = [
        {'key': 'in_progress', 'title': '进行中', 'plans': [p for p in main_plans if getattr(p, 'ui_status_key', '') == 'in_progress']},
        {'key': 'completed', 'title': '已完成', 'plans': [p for p in main_plans if getattr(p, 'ui_status_key', '') == 'completed']},
        {'key': 'delayed', 'title': '延期', 'plans': [p for p in main_plans if getattr(p, 'ui_status_key', '') == 'delayed']},
    ]
    
    context = {
        'plans': main_plans,
        'all_plans': all_plans,
        'members_with_plans': members_with_plans,
        'selected_member_data': selected_member_data,
        'total_plan_count': total_plan_count,
        'total_completed_count': total_completed_count,
        'total_in_progress_count': total_in_progress_count,
        'avg_completion_rate': avg_completion_rate,
        'delayed_plan_count': delayed_plan_count,
        'view': view,
        'selected_user': selected_user,
        'requested_member': requested_member,
        'member_query': member_query,
        'plan_query': plan_query,
        'status_filter': status_filter,
        'board_columns': board_columns,
    }
    
    if view == 'calendar':
        # 获取当前月份
        month_param = request.GET.get('month')
        if month_param:
            try:
                year, month = map(int, month_param.split('-'))
            except:
                today = datetime.now()
                year, month = today.year, today.month
        else:
            today = datetime.now()
            year, month = today.year, today.month
        
        # 计算上个月和下个月
        if month == 1:
            prev_month = f"{year-1}-12"
            next_month = f"{year}-2"
        elif month == 12:
            prev_month = f"{year}-{month-1}"
            next_month = f"{year+1}-1"
        else:
            prev_month = f"{year}-{month-1}"
            next_month = f"{year}-{month+1}"
        
        # 获取当月天数
        month_days = cal.monthrange(year, month)[1]
        
        # 获取当月所有计划（包括子计划）
        month_plans = all_plans.filter(date__year=year, date__month=month)
        
        # 构建日历数据
        calendar_days = []
        first_day = datetime(year, month, 1).weekday()
        
        # 添加空白的起始天
        for _ in range(first_day):
            calendar_days.append({'day': '', 'plans': [], 'is_today': False})
        
        today = datetime.now()
        for day in range(1, month_days + 1):
            day_plans = month_plans.filter(date=datetime(year, month, day).date())
            calendar_days.append({
                'day': day,
                'plans': list(day_plans),
                'is_today': (today.year == year and today.month == month and today.day == day)
            })
        
        context.update({
            'year': year,
            'month': month,
            'prev_month': prev_month,
            'next_month': next_month,
            'calendar_days': calendar_days
        })
    
    return render(request, 'lab_management/plan_list.html', context)


def team_overview(request):
    members = UserProfile.objects.select_related('user').filter(user__is_superuser=False).all()
    today = timezone.localdate()
    period_start = today.replace(day=1)
    # 只查询公开的计划
    all_plans = DailyPlan.objects.select_related('user__profile').filter(user__is_superuser=False, parent__isnull=True, is_public=True).prefetch_related('sub_plans').order_by('-date')
    all_achievements = Achievement.objects.filter(is_public=True, user__is_superuser=False).select_related('user__profile').all()
    public_documents = Document.objects.filter(is_public=True, user__is_superuser=False).select_related('user')
    
    member_data = []
    members_with_plans = []
    total_completed_all = 0
    total_count_all = 0
    
    for profile in members:
        user_plans = list(all_plans.filter(user=profile.user))
        
        # 团队总览只统计主计划，子计划作为计划详情，不重复计入总数。
        main_completed = len([p for p in user_plans if p.is_completed])
        main_total = len(user_plans)

        total_completed = main_completed
        total_count = main_total
        in_progress_count = max(total_count - total_completed, 0)
        total_completed_all += total_completed
        total_count_all += total_count
        
        user_documents = public_documents.filter(user=profile.user)
        user_achievements = all_achievements.filter(user=profile.user)
        update_dates = [p.date for p in user_plans if p.date]
        update_dates += [doc.uploaded_at.date() for doc in user_documents if doc.uploaded_at]
        update_dates += [(ach.date or ach.created_at.date()) for ach in user_achievements if ach.date or ach.created_at]
        latest_update = max(update_dates) if update_dates else profile.user.date_joined.date()
        
        completion_rate = round(total_completed / total_count * 100, 0) if total_count > 0 else 0
        
        member_data.append({
            'user': profile.user,
            'profile': profile,
            'plans': user_plans[:5],
            'completed_count': total_completed,
            'in_progress_count': in_progress_count,
            'total_count': total_count,
            'documents': user_documents,
            'achievements': user_achievements,
            'completion_rate': completion_rate,
            'latest_update': latest_update,
        })
        
        members_with_plans.append({
            'profile': profile,
            'plans': user_plans,
            'plan_count': total_count,
            'completed_count': total_completed,
            'in_progress_count': in_progress_count,
            'completion_rate': completion_rate,
            'latest_update': latest_update,
        })

    member_data.sort(key=lambda item: (item['latest_update'], item['completion_rate'], item['total_count']), reverse=True)
    current_year_achievements = all_achievements.filter(
        models.Q(date__year=today.year) |
        models.Q(date__isnull=True, created_at__year=today.year)
    )
    achievement_counts = {
        'paper': current_year_achievements.filter(achievement_type='paper').count(),
        'patent': current_year_achievements.filter(achievement_type='patent').count(),
        'project': current_year_achievements.filter(achievement_type='project').count(),
        'award': current_year_achievements.filter(achievement_type='award').count(),
    }
    member_type_counts = {
        'phd': members.filter(member_type='phd').count(),
        'master': members.filter(member_type='master').count(),
        'undergraduate': members.filter(member_type='undergraduate').count(),
        'graduated': members.filter(member_type='graduated').count(),
    }
    monthly_document_count = public_documents.filter(uploaded_at__date__gte=period_start, uploaded_at__date__lte=today).count()
    avg_completion_rate = round(total_completed_all / total_count_all * 100, 0) if total_count_all > 0 else 0
    weekday_names = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    
    context = {
        'member_data': member_data,
        'members_with_plans': members_with_plans,
        'all_plans': all_plans[:20],
        'all_achievements': all_achievements[:30],
        'total_plan_count': total_count_all,
        'total_completed_count': total_completed_all,
        'avg_completion_rate': avg_completion_rate,
        'monthly_document_count': monthly_document_count,
        'achievement_counts': achievement_counts,
        'member_type_counts': member_type_counts,
        'current_date': today,
        'current_weekday': weekday_names[today.weekday()],
        'period_start': period_start,
        'period_end': today,
    }
    return render(request, 'lab_management/team_overview.html', context)


def user_profile(request, username):
    user = get_object_or_404(User, username=username)
    profile = get_object_or_404(UserProfile, user=user)

    can_view_all = request.user.is_authenticated and (request.user.id == user.id or request.user.is_superuser)

    documents_qs = Document.objects.select_related('user', 'user__profile').filter(user=user)
    if not can_view_all:
        documents_qs = documents_qs.filter(is_public=True)
    documents = list(documents_qs.order_by('-uploaded_at')[:10])

    if can_view_all:
        plans_qs = DailyPlan.objects.filter(user=user, parent__isnull=True)
    else:
        plans_qs = DailyPlan.objects.filter(user=user, parent__isnull=True, is_public=True)
    plans = list(plans_qs.prefetch_related('sub_plans').order_by('-date', '-id')[:10])

    for plan in plans:
        sub_plan_qs = plan.sub_plans.all() if can_view_all else plan.sub_plans.filter(is_public=True)
        visible_sub_plans = list(sub_plan_qs.order_by('date', 'order', 'id'))
        completed_count = sum(1 for sub_plan in visible_sub_plans if sub_plan.is_completed)
        total_count = len(visible_sub_plans)
        if total_count:
            progress_rate = round(completed_count / total_count * 100)
        else:
            progress_rate = 100 if plan.is_completed else 0
        plan.visible_sub_plans = visible_sub_plans
        plan.sub_total = total_count
        plan.sub_completed = completed_count
        plan.progress_rate = progress_rate

    achievements_qs = Achievement.objects.filter(user=user)
    if not can_view_all:
        achievements_qs = achievements_qs.filter(is_public=True)
    achievements = list(achievements_qs.order_by('-date', '-created_at')[:10])
    messages_list = Message.objects.select_related('sender__profile').filter(receiver=user).order_by('-created_at')[:20]

    context = {
        'profile_user': user,
        'profile': profile,
        'documents': documents,
        'plans': plans,
        'achievements': achievements,
        'messages_list': messages_list,
        'can_view_all': can_view_all,
        'document_count': documents_qs.count(),
        'plan_count': plans_qs.count(),
        'achievement_count': achievements_qs.count(),
        'message_count': Message.objects.filter(receiver=user).count(),
    }
    return render(request, 'lab_management/user_profile.html', context)


def document_list(request):
    from datetime import timedelta
    from django.core.paginator import Paginator
    query = request.GET.get('q', '').strip()
    doc_type = request.GET.get('type', 'all')
    uploader = request.GET.get('uploader', 'all')
    scope = request.GET.get('scope', 'all')
    category = request.GET.get('category', 'all')
    sort = request.GET.get('sort', 'latest')
    view = request.GET.get('view', 'list')
    valid_types = {'all', 'pdf', 'word', 'excel', 'ppt', 'txt', 'zip', 'image', 'data', 'code', 'other'}
    valid_scopes = {'all', 'public', 'private'}
    valid_categories = {'all', 'project', 'paper', 'experiment', 'learning', 'meeting'}
    valid_sorts = {'latest', 'downloads', 'name'}
    if doc_type not in valid_types:
        doc_type = 'all'
    if scope not in valid_scopes:
        scope = 'all'
    if category not in valid_categories:
        category = 'all'
    if sort not in valid_sorts:
        sort = 'latest'
    if view not in {'list', 'grid'}:
        view = 'list'

    base_documents = Document.objects.select_related('user', 'user__profile')
    if request.user.is_authenticated:
        if not request.user.is_superuser:
            base_documents = base_documents.filter(models.Q(is_public=True) | models.Q(user=request.user))
    else:
        base_documents = base_documents.filter(is_public=True)

    documents = base_documents
    if query:
        documents = documents.filter(
            models.Q(title__icontains=query) |
            models.Q(description__icontains=query) |
            models.Q(user__username__icontains=query) |
            models.Q(user__first_name__icontains=query)
        )
    if category != 'all':
        documents = documents.filter(category=category)
    if uploader != 'all':
        documents = documents.filter(user__username=uploader)
    if scope == 'public':
        documents = documents.filter(is_public=True)
    elif scope == 'private':
        documents = documents.filter(is_public=False)
    if doc_type != 'all':
        if doc_type in {'pdf', 'word', 'excel', 'txt', 'other'}:
            documents = documents.filter(document_type=doc_type)
        else:
            ext_map = {
                'ppt': ['.ppt', '.pptx'],
                'zip': ['.zip', '.rar', '.7z'],
                'image': ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp'],
                'data': ['.csv', '.json'],
                'code': ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c'],
            }
            q_obj = models.Q()
            for ext in ext_map.get(doc_type, []):
                q_obj |= models.Q(file__iendswith=ext)
            documents = documents.filter(q_obj)

    if sort == 'downloads':
        documents = documents.order_by('-download_count', '-uploaded_at')
    elif sort == 'name':
        documents = documents.order_by('title', '-uploaded_at')
    else:
        documents = documents.order_by('-uploaded_at')

    recent_since = timezone.now() - timedelta(days=7)
    document_count = base_documents.count()
    public_document_count = base_documents.filter(is_public=True).count()
    recent_upload_count = base_documents.filter(uploaded_at__gte=recent_since).count()
    total_download_count = base_documents.aggregate(total=models.Sum('download_count'))['total'] or 0
    storage_bytes = 0
    for doc in base_documents:
        try:
            if doc.file:
                storage_bytes += doc.file.size
        except Exception:
            pass
    storage_limit_bytes = 100 * 1024 * 1024 * 1024
    storage_used_gb = round(storage_bytes / (1024 * 1024 * 1024), 2)
    storage_percent = round(storage_bytes / storage_limit_bytes * 100, 1) if storage_limit_bytes else 0
    uploader_options = User.objects.filter(documents__in=base_documents).distinct().order_by('first_name', 'username')
    paginator = Paginator(documents, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    for doc in page_obj.object_list:
        ext = doc.get_file_ext()
        if ext in ['.ppt', '.pptx']:
            doc.ui_type = 'PPT'
            doc.ui_icon = 'ppt.png'
        elif ext in ['.zip', '.rar', '.7z']:
            doc.ui_type = '压缩包'
            doc.ui_icon = 'zip.png'
        elif ext in ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp']:
            doc.ui_type = '图片'
            doc.ui_icon = 'image.png'
        elif ext in ['.csv', '.json']:
            doc.ui_type = '数据'
            doc.ui_icon = 'data.png'
        elif ext in ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c']:
            doc.ui_type = '代码'
            doc.ui_icon = 'code.svg'
        elif doc.is_pdf():
            doc.ui_type = 'PDF'
            doc.ui_icon = 'pdf.png'
        elif doc.is_word():
            doc.ui_type = 'Word'
            doc.ui_icon = 'word.png'
        elif doc.is_excel():
            doc.ui_type = 'Excel'
            doc.ui_icon = 'excel.png'
        elif doc.is_txt():
            doc.ui_type = '文本'
            doc.ui_icon = 'txt.png'
        else:
            doc.ui_type = doc.get_document_type_display()
            doc.ui_icon = 'folder.svg'

    context = {
        'documents': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'document_count': document_count,
        'filtered_count': paginator.count,
        'recent_upload_count': recent_upload_count,
        'public_document_count': public_document_count,
        'public_ratio': round(public_document_count / document_count * 100, 1) if document_count else 0,
        'total_download_count': total_download_count,
        'storage_used_gb': storage_used_gb,
        'storage_percent': storage_percent,
        'uploader_options': uploader_options,
        'query': query,
        'doc_type': doc_type,
        'uploader': uploader,
        'scope': scope,
        'category': category,
        'sort': sort,
        'view': view,
    }
    return render(request, 'lab_management/document_list.html', context)


def announcement_list(request):
    category = request.GET.get('category', 'all')
    query = request.GET.get('q', '').strip()
    page_size = request.GET.get('page_size', '10')
    if category not in {'all', 'system', 'academic', 'team', 'important'}:
        category = 'all'
    if page_size not in {'5', '10', '20'}:
        page_size = '10'

    base_announcements = Announcement.objects.select_related('author').annotate(
        comment_count=models.Count('comments')
    ).order_by('-is_pinned', '-created_at', '-id')
    announcements = base_announcements
    if category != 'all':
        announcements = announcements.filter(category=category)
    if query:
        announcements = announcements.filter(
            models.Q(title__icontains=query) |
            models.Q(content__icontains=query) |
            models.Q(author__username__icontains=query) |
            models.Q(author__first_name__icontains=query)
        )

    paginator = Paginator(announcements, int(page_size))
    page_obj = paginator.get_page(request.GET.get('page', 1))

    category_ui = {
        'system': ('系统公告', 'system', 'megaphone'),
        'academic': ('学术通知', 'academic', 'cap'),
        'team': ('团队动态', 'team', 'users'),
        'important': ('重要通知', 'important', 'file'),
    }
    for announcement in page_obj.object_list:
        announcement.comments_list = announcement.comments.select_related('author')[:10]
        label, tone, icon = category_ui.get(announcement.category, category_ui['system'])
        announcement.ui_category = label
        announcement.ui_tone = tone
        announcement.ui_icon = icon
    
    # 检查用户是否有发布公告的权限
    can_post_announcement = False
    if request.user.is_authenticated:
        can_post_announcement = request.user.is_superuser or (
            hasattr(request.user, 'profile') and request.user.profile.can_post_announcement
        )
    
    context = {
        'announcements': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'announcement_total': announcements.count(),
        'all_count': base_announcements.count(),
        'system_count': base_announcements.filter(category='system').count(),
        'academic_count': base_announcements.filter(category='academic').count(),
        'team_count': base_announcements.filter(category='team').count(),
        'important_count': base_announcements.filter(category='important').count(),
        'category': category,
        'query': query,
        'page_size': page_size,
        'can_post_announcement': can_post_announcement,
    }
    return render(request, 'lab_management/announcement_list.html', context)


@login_required
def add_announcement_comment(request, ann_id):
    announcement = get_object_or_404(Announcement, id=ann_id)
    
    if not announcement.comments_enabled:
        messages.error(request, '该公告已关闭评论功能')
        return redirect('announcement_list')
    
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if content:
            AnnouncementComment.objects.create(
                announcement=announcement,
                author=request.user,
                content=content
            )
            messages.success(request, '评论成功！')
        else:
            messages.error(request, '评论内容不能为空')
    
    return redirect('announcement_list')


@login_required
def delete_announcement_comment(request, comment_id):
    comment = get_object_or_404(AnnouncementComment, id=comment_id)
    
    if request.user == comment.author or request.user.is_superuser:
        comment.delete()
        messages.success(request, '评论已删除')
    else:
        messages.error(request, '您没有权限删除此评论')
    
    return redirect('announcement_list')


@login_required
def my_page(request):
    user = request.user
    active_tab = request.GET.get('tab', 'profile')
    valid_tabs = {'profile', 'plans', 'documents', 'achievements', 'management'}
    if active_tab not in valid_tabs:
        active_tab = 'profile'
    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        UserProfile.objects.create(user=user, research_topic='待填写')
        profile = user.profile
    if active_tab == 'management' and not (user.is_superuser or profile.can_post_announcement):
        active_tab = 'profile'
    
    documents = Document.objects.filter(user=user).order_by('-uploaded_at')
    plans = DailyPlan.objects.filter(user=user, parent__isnull=True).prefetch_related('sub_plans').order_by('-date', '-id')
    achievements = Achievement.objects.filter(user=user).order_by('-date', '-created_at')
    received_messages = Message.objects.select_related('sender__profile').filter(receiver=user).order_by('-created_at')[:20]
    recent_announcements = Announcement.objects.select_related('author').annotate(
        comment_count=models.Count('comments')
    ).order_by('-is_pinned', '-created_at')[:4]
    completed_plans_count = plans.filter(is_completed=True).count()
    in_progress_plans_count = plans.filter(is_completed=False).count()
    public_documents_count = documents.filter(is_public=True).count()
    private_documents_count = documents.filter(is_public=False).count()
    public_achievements_count = achievements.filter(is_public=True).count()
    unread_message_count = Message.objects.filter(receiver=user, is_read=False).count()

    for plan in plans:
        sub_plans = list(plan.sub_plans.all())
        sub_total = len(sub_plans)
        sub_completed = len([sub for sub in sub_plans if sub.is_completed])
        plan.workspace_progress = 100 if plan.is_completed else (round(sub_completed / sub_total * 100) if sub_total else 0)
        plan.workspace_sub_completed = sub_completed
        plan.workspace_sub_total = sub_total
        
        # 正确的状态判断逻辑
        if plan.is_completed:
            plan.workspace_status = '已完成'
        elif plan.is_delayed:
            plan.workspace_status = '已延期'
        elif plan.deadline:
            if plan.deadline < timezone.localdate():
                plan.workspace_status = '已超期'
            else:
                plan.workspace_status = '进行中'
        else:
            plan.workspace_status = '进行中'
        
        # 截止时间显示（deadline 或 "至今"）
        if plan.deadline:
            plan.workspace_deadline_display = plan.deadline.strftime('%m-%d')
        else:
            plan.workspace_deadline_display = '至今'
    
    # 标记收到的消息为已读
    Message.objects.filter(receiver=user, is_read=False).update(is_read=True)
    
    context = {
        'active_tab': active_tab,
        'profile': profile,
        'documents': documents,
        'plans': plans,
        'achievements': achievements,
        'received_messages': received_messages,
        'recent_announcements': recent_announcements,
        'completed_plans_count': completed_plans_count,
        'today': timezone.localdate(),
        'in_progress_plans_count': in_progress_plans_count,
        'public_documents_count': public_documents_count,
        'private_documents_count': private_documents_count,
        'public_achievements_count': public_achievements_count,
        'unread_message_count': unread_message_count,
    }
    return render(request, 'lab_management/my_page.html', context)


@login_required
def add_document(request):
    if request.method == 'POST':
        next_url = request.POST.get('next') or 'my_page'
        title = request.POST.get('title')
        description = request.POST.get('description')
        file = request.FILES.get('file')
        
        if not file:
            messages.error(request, '请选择要上传的文件')
            return redirect(next_url)
        
        ext = file.name.split('.')[-1].lower()
        doc_type = 'other'
        if ext in ['pdf']:
            doc_type = 'pdf'
        elif ext in ['doc', 'docx']:
            doc_type = 'word'
        elif ext in ['xls', 'xlsx']:
            doc_type = 'excel'
        elif ext in ['txt']:
            doc_type = 'txt'
        
        is_public = request.POST.get('is_public') == 'on'
        category = request.POST.get('category', 'project')
        valid_categories = {choice[0] for choice in Document.CATEGORY_CHOICES}
        if category not in valid_categories:
            category = 'project'
        
        Document.objects.create(
            user=request.user,
            title=title,
            file=file,
            document_type=doc_type,
            category=category,
            description=description,
            is_public=is_public
        )
        
        messages.success(request, '文档上传成功！')
        return redirect(next_url)
    
    return redirect('my_page')


@login_required
def download_document(request, doc_id):
    """下载文档（强制下载而不是预览）"""
    try:
        doc = Document.objects.get(id=doc_id)
        
        # 检查权限：私密文档只有所有者或超级用户可以下载
        if not doc.is_public:
            if request.user != doc.user and not request.user.is_superuser:
                messages.error(request, '您没有权限下载此文档')
                return redirect('document_list')
        
        if not doc.file or not os.path.exists(doc.file.path):
            messages.error(request, '文件不存在或已被删除')
            return redirect('document_list')
        Document.objects.filter(id=doc.id).update(download_count=models.F('download_count') + 1)
        
        # 读取文件
        with open(doc.file.path, 'rb') as f:
            file_content = f.read()
        
        # 使用文档标题作为下载文件名，保留原始扩展名
        import os.path as op
        # 获取原始文件的扩展名
        original_ext = op.splitext(doc.file.name)[1]  # 如 .txt, .pdf
        # 使用文档标题 + 原始扩展名
        download_filename = f"{doc.title}{original_ext}"
        # 处理文件名中的特殊字符
        download_filename = download_filename.replace('/', '_').replace('\\', '_')
        
        # 创建下载响应
        response = HttpResponse(file_content, content_type='application/octet-stream')
        # 使用 RFC 5987 编码支持中文文件名
        from urllib.parse import quote
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(download_filename)}"
        response['Content-Length'] = len(file_content)
        
        return response
        
    except Document.DoesNotExist:
        messages.error(request, '文档不存在')
        return redirect('document_list')
    except Exception as e:
        messages.error(request, f'下载失败')
        return redirect('document_list')


def preview_document(request, doc_id):
    """预览文档"""
    try:
        doc = Document.objects.get(id=doc_id)
        
        # 检查权限
        if not doc.is_public:
            if not request.user.is_authenticated or (request.user != doc.user and not request.user.is_superuser):
                return HttpResponse("无权限", status=403)
        
        if not doc.file or not os.path.exists(doc.file.path):
            messages.error(request, '文件不存在或已被删除')
            return redirect('document_list')
        
        # 获取文件扩展名
        file_ext = doc.file.name.split('.')[-1].lower()
        
        # 支持预览的文件类型
        preview_types = ['pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'txt']
        
        if file_ext in ['doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'pdf']:
            # Office 文档和 PDF 暂不支持预览
            html = '<html><head><meta charset="utf-8"></head><body style="font-family: sans-serif; padding: 40px; text-align: center;"><h2>暂不支持在线预览</h2><p>该文件类型暂不支持在线预览，请下载后查看</p><a href="/documents/" style="color: #2563eb;">返回文档列表</a></body></html>'
            return HttpResponse(html, content_type='text/html; charset=utf-8')
        
        if file_ext == 'txt':
            # TXT 文件直接读取内容
            try:
                with open(doc.file.path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                return HttpResponse(f'<pre style="padding: 20px; font-family: monospace; white-space: pre-wrap; word-wrap: break-word;">{content}</pre>', content_type='text/html; charset=utf-8')
            except:
                return download_document(request, doc_id)
        
        # 图片文件直接显示
        if file_ext in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
            file_url = doc.file.url
            html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>预览 - {doc.title}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #1f2937; min-height: 100vh; }}
        .header {{
            background: rgba(0,0,0,0.5);
            padding: 12px 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            z-index: 100;
        }}
        .header h1 {{ font-size: 16px; color: white; font-weight: 600; }}
        .header .actions {{ display: flex; gap: 10px; }}
        .header a {{
            padding: 6px 14px;
            border-radius: 6px;
            text-decoration: none;
            font-size: 13px;
            font-weight: 500;
        }}
        .btn-primary {{ background: #6366f1; color: white; }}
        .btn-secondary {{ background: rgba(255,255,255,0.2); color: white; }}
        .preview-container {{
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
            padding: 80px 20px 20px;
        }}
        .preview-container img {{
            max-width: 100%;
            max-height: calc(100vh - 100px);
            object-fit: contain;
            border-radius: 8px;
            box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.3);
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🖼️ {doc.title}</h1>
        <div class="actions">
            <a href="{file_url}" download class="btn-primary">下载</a>
            <a href="/documents/" class="btn-secondary">返回</a>
        </div>
    </div>
    <div class="preview-container">
        <img src="{file_url}" alt="{doc.title}">
    </div>
</body>
</html>'''
            return HttpResponse(html, content_type='text/html; charset=utf-8')
        
        # 其他不支持预览的类型
        html = '<html><head><meta charset="utf-8"></head><body style="font-family: sans-serif; padding: 40px; text-align: center;"><h2>暂不支持在线预览</h2><p>该文件类型暂不支持在线预览，请下载后查看</p><a href="/documents/" style="color: #2563eb;">返回文档列表</a></body></html>'
        return HttpResponse(html, content_type='text/html; charset=utf-8')
        
    except Document.DoesNotExist:
        messages.error(request, '文档不存在')
        return redirect('document_list')
    except Exception as e:
        messages.error(request, f'预览失败')
        return redirect('document_list')


@login_required
def delete_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    if request.user == doc.user or request.user.is_superuser:
        doc.file.delete()
        doc.delete()
        messages.success(request, '文档删除成功！')
    else:
        messages.error(request, '您没有权限删除此文档')
    return redirect('my_page')


@login_required
def add_plan(request):
    if request.method == 'POST':
        next_url = request.POST.get('next') or 'my_page'
        content = request.POST.get('content')
        date = request.POST.get('date')
        deadline = request.POST.get('deadline')
        parent_id = request.POST.get('parent_id')
        
        if not content:
            messages.error(request, '请填写计划内容')
            return redirect(next_url)
        
        # 如果有 parent_id，则是添加子计划
        parent_plan = None
        if parent_id:
            parent_plan = get_object_or_404(DailyPlan, id=parent_id)
            # 检查权限
            if request.user != parent_plan.user and not request.user.is_superuser:
                messages.error(request, '您没有权限为此计划添加子计划')
                return redirect('plan_list')
        
        DailyPlan.objects.create(
            user=request.user,
            content=content,
            date=date or None,
            deadline=deadline or None,  # 截止时间（可选）
            parent=parent_plan  # 设置父计划
        )
        
        if parent_plan:
            messages.success(request, '子计划添加成功！')
            return redirect('plan_list')
        
        messages.success(request, '计划添加成功！')
        return redirect(next_url)
    
    return redirect('my_page')


@login_required
def edit_plan(request, plan_id):
    plan = get_object_or_404(DailyPlan, id=plan_id)
    
    if request.user != plan.user and not request.user.is_superuser:
        messages.error(request, '您没有权限修改此计划')
        return redirect('my_page')
    
    if request.method == 'POST':
        plan.content = request.POST.get('content')
        plan.date = request.POST.get('date') or None
        plan.deadline = request.POST.get('deadline') or None
        plan.is_completed = request.POST.get('is_completed') == 'on'
        plan.is_public = request.POST.get('is_public') == 'on'
        plan.save()
        messages.success(request, '计划更新成功！')
        return redirect('my_page')
    
    context = {'plan': plan}
    return render(request, 'lab_management/edit_plan.html', context)


@login_required
def delete_plan(request, plan_id):
    plan = get_object_or_404(DailyPlan, id=plan_id)
    if request.user == plan.user or request.user.is_superuser:
        plan.delete()
        messages.success(request, '计划删除成功！')
    else:
        messages.error(request, '您没有权限删除此计划')
    return redirect('my_page')


@login_required
def toggle_plan(request, plan_id):
    plan = get_object_or_404(DailyPlan, id=plan_id)
    if request.user == plan.user or request.user.is_superuser:
        plan.is_completed = not plan.is_completed
        plan.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': '没有权限'})


@login_required
@require_POST
def extend_deadline(request, plan_id):
    """延期计划截止时间"""
    plan = get_object_or_404(DailyPlan, id=plan_id)
    
    if request.user != plan.user and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': '没有权限'})
    
    new_deadline = request.POST.get('deadline')
    if not new_deadline:
        return JsonResponse({'success': False, 'error': '请选择新的截止时间'})
    
    plan.deadline = new_deadline
    plan.is_delayed = True
    plan.save()
    
    return JsonResponse({
        'success': True, 
        'message': '计划已延期',
        'new_deadline': new_deadline
    })


@login_required
@require_POST
def add_sub_plan_ajax(request):
    """AJAX 添加子计划"""
    content = request.POST.get('content')
    date = request.POST.get('date')
    parent_id = request.POST.get('parent_id')
    
    if not content:
        return JsonResponse({'success': False, 'error': '请填写计划内容'})
    
    if not parent_id:
        return JsonResponse({'success': False, 'error': '缺少主计划ID'})
    
    parent_plan = get_object_or_404(DailyPlan, id=parent_id)
    
    # 检查权限
    if request.user != parent_plan.user and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': '您没有权限为此计划添加子计划'})
    
    # 创建子计划
    sub_plan = DailyPlan.objects.create(
        user=request.user,
        content=content,
        date=date or timezone.localdate(),
        parent=parent_plan
    )
    
    # 返回子计划数据
    parent_end_date = parent_plan.end_date
    return JsonResponse({
        'success': True,
        'sub_plan': {
            'id': sub_plan.id,
            'content': sub_plan.content,
            'date': str(sub_plan.date),
            'is_completed': sub_plan.is_completed,
            'parent_id': parent_plan.id,
            'parent_date_range': parent_plan.date_range,
            'parent_end_date': str(parent_end_date) if parent_end_date else None,
        }
    })


@login_required
def add_announcement(request):
    # 检查用户是否有发布公告的权限
    can_post = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.can_post_announcement
    )
    if not can_post:
        messages.error(request, '您没有发布公告的权限')
        return redirect('announcement_list')
    
    if request.method == 'POST':
        next_url = request.POST.get('next') or 'announcement_list'
        title = request.POST.get('title')
        content = request.POST.get('content')
        category = request.POST.get('category', 'system')
        if category not in {'system', 'academic', 'team', 'important'}:
            category = 'system'
        is_pinned = request.user.is_superuser and request.POST.get('is_pinned') == 'on'
        comments_enabled = request.POST.get('comments_enabled') != 'off'
        
        if not title or not content:
            messages.error(request, '请填写标题和内容')
            return redirect(next_url)
        
        Announcement.objects.create(
            author=request.user,
            title=title,
            content=content,
            category=category,
            is_pinned=is_pinned,
            comments_enabled=comments_enabled
        )
        
        messages.success(request, '公告发布成功！')
        return redirect(next_url)
    
    # GET 请求显示发布表单
    return render(request, 'lab_management/create_announcement.html')


@login_required
def edit_announcement(request, ann_id):
    announcement = get_object_or_404(Announcement, id=ann_id)
    
    if request.user != announcement.author and not request.user.is_superuser:
        messages.error(request, '您没有权限修改此公告')
        return redirect('my_page')
    
    if request.method == 'POST':
        announcement.title = request.POST.get('title')
        announcement.content = request.POST.get('content')
        category = request.POST.get('category', announcement.category)
        if category in {'system', 'academic', 'team', 'important'}:
            announcement.category = category
        if request.user.is_superuser:
            announcement.is_pinned = request.POST.get('is_pinned') == 'on'
        if request.user.is_superuser or request.user == announcement.author:
            announcement.comments_enabled = request.POST.get('comments_enabled') == 'on'
        announcement.save()
        messages.success(request, '公告更新成功！')
        return redirect('announcement_list')
    
    context = {'announcement': announcement}
    return render(request, 'lab_management/edit_announcement.html', context)


@login_required
def delete_announcement(request, ann_id):
    announcement = get_object_or_404(Announcement, id=ann_id)
    if request.user == announcement.author or request.user.is_superuser:
        announcement.delete()
        messages.success(request, '公告删除成功！')
    else:
        messages.error(request, '您没有权限删除此公告')
    return redirect('announcement_list')


@login_required
def edit_profile(request):
    if request.method == 'POST':
        next_url = request.POST.get('next') or 'my_page'
        profile = request.user.profile
        profile.research_topic = request.POST.get('research_topic', '')
        profile.bio = request.POST.get('bio', '')
        profile.personal_page = request.POST.get('personal_page', '')
        profile.email = request.POST.get('email', '')
        profile.phone = request.POST.get('phone', '')
        
        # 更新成员分类信息
        member_type = request.POST.get('member_type', 'master')
        grade = request.POST.get('grade', '1')
        valid_member_types = {choice[0] for choice in UserProfile.MEMBER_TYPE_CHOICES}
        valid_grades = {choice[0] for choice in UserProfile.GRADE_CHOICES}
        profile.member_type = member_type if member_type in valid_member_types else 'master'
        profile.grade = grade if grade in valid_grades else '1'
        
        # 处理头像上传
        if request.FILES.get('avatar'):
            profile.avatar = request.FILES.get('avatar')
        
        profile.save()
        
        user = request.user
        user.first_name = request.POST.get('first_name', '')
        user.save()
        
        messages.success(request, '资料更新成功！')
        return redirect(next_url)
    
    return redirect('my_page')


@login_required
def add_achievement(request):
    if request.method == 'POST':
        next_url = request.POST.get('next') or 'my_page'
        title = request.POST.get('title')
        achievement_type = request.POST.get('achievement_type')
        description = request.POST.get('description', '')
        link = request.POST.get('link', '')
        date = request.POST.get('date')
        is_public = request.POST.get('is_public', 'on') == 'on'
        if not title:
            messages.error(request, '请填写成果名称')
            return redirect(next_url)
        
        achievement = Achievement.objects.create(
            user=request.user,
            title=title,
            achievement_type=achievement_type,
            description=description,
            link=link,
            date=date or None,
            is_public=is_public,
        )
        
        if request.FILES.get('file'):
            achievement.file = request.FILES.get('file')
            achievement.save()
        
        messages.success(request, '成果添加成功！')
        return redirect(next_url)
    
    return redirect('my_page')


@login_required
def delete_achievement(request, ach_id):
    achievement = get_object_or_404(Achievement, id=ach_id)
    if request.user == achievement.user or request.user.is_superuser:
        if achievement.file:
            achievement.file.delete()
        achievement.delete()
        messages.success(request, '成果已删除')
    else:
        messages.error(request, '您没有权限删除此成果')
    return redirect('my_page')


@login_required
def send_message(request, username):
    receiver = get_object_or_404(User, username=username)
    if request.method == 'POST':
        content = request.POST.get('content')
        if content:
            Message.objects.create(
                sender=request.user,
                receiver=receiver,
                content=content,
            )
            messages.success(request, f'留言已发送给 {receiver.first_name or receiver.username}')
        else:
            messages.error(request, '留言内容不能为空')
    return redirect('user_profile', username=username)


@login_required
def delete_message(request, msg_id):
    message = get_object_or_404(Message, id=msg_id)
    if request.user == message.receiver or request.user == message.sender:
        message.delete()
        messages.success(request, '留言已删除')
    else:
        messages.error(request, '您没有权限删除此留言')
    
    if request.user == message.receiver:
        return redirect('my_page')
    else:
        return redirect('user_profile', username=message.receiver.username)


@login_required
def delete_user(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, '只有管理员可以删除用户')
        return redirect('my_page')
    
    user_to_delete = get_object_or_404(User, id=user_id)
    if user_to_delete == request.user:
        messages.error(request, '不能删除自己的账号')
        return redirect('my_page')
    
    for profile in user_to_delete.profile.documents.all():
        profile.file.delete()
    user_to_delete.delete()
    messages.success(request, '用户已删除')
    return redirect('home')


@login_required
def search(request):
    """站内搜索"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return render(request, 'lab_management/search.html', {
            'query': '',
            'results': {'members': [], 'achievements': [], 'documents': [], 'announcements': [], 'plans': []}
        })
    
    from .models import UserProfile, Achievement, Document, Announcement, DailyPlan
    
    results = {
        'members': [],
        'achievements': [],
        'documents': [],
        'announcements': [],
        'plans': []
    }
    
    # 搜索成员（用户名、姓名、研究方向）
    profiles = UserProfile.objects.select_related('user').filter(
        user__is_superuser=False
    ).filter(
        models.Q(user__username__icontains=query) |
        models.Q(user__first_name__icontains=query) |
        models.Q(research_topic__icontains=query) |
        models.Q(bio__icontains=query)
    )[:10]
    results['members'] = profiles
    
    # 搜索成果
    achievements = Achievement.objects.filter(
        models.Q(title__icontains=query) |
        models.Q(description__icontains=query) |
        models.Q(achievement_type__icontains=query)
    ).select_related('user')[:10]
    results['achievements'] = achievements
    
    # 搜索文档（只显示公开的）
    documents = Document.objects.filter(
        is_public=True
    ).filter(
        models.Q(title__icontains=query) |
        models.Q(description__icontains=query)
    ).select_related('user')[:10]
    results['documents'] = documents
    
    # 搜索公告
    announcements = Announcement.objects.filter(
        models.Q(title__icontains=query) |
        models.Q(content__icontains=query)
    ).select_related('author')[:10]
    results['announcements'] = announcements
    
    # 搜索工作计划
    plans = DailyPlan.objects.filter(
        models.Q(content__icontains=query)
    ).select_related('user')[:10]
    results['plans'] = plans
    
    return render(request, 'lab_management/search.html', {
        'query': query,
        'results': results
    })


def get_lab_data_for_ai(question):
    """根据问题获取相关的实验室数据"""
    from .models import DailyPlan, UserProfile, Achievement, Document, Announcement
    from django.contrib.auth.models import User
    from django.utils import timezone
    
    data_context = []
    question_lower = question.lower()
    
    # 检查是否询问任务/计划相关
    if any(keyword in question_lower for keyword in ['任务', '计划', '工作', '完成', '进度', 'todo']):
        plans = DailyPlan.objects.all().select_related('user__profile').order_by('-date', '-id')[:20]
        if plans:
            data_context.append("【工作计划列表】")
            completed_count = {}
            pending_count = {}
            for plan in plans:
                username = plan.user.username
                if plan.is_completed:
                    completed_count[username] = completed_count.get(username, 0) + 1
                else:
                    pending_count[username] = pending_count.get(username, 0) + 1
                status_text = "✅已完成" if plan.is_completed else "⏳进行中"
                data_context.append(f"- {plan.user.username}: {plan.content} ({status_text}, {plan.date})")
            
            if completed_count:
                data_context.append("\n【任务完成统计】")
                for username, count in sorted(completed_count.items(), key=lambda x: x[1], reverse=True):
                    data_context.append(f"- {username}: 完成 {count} 个任务")
        
    # 检查是否询问成员相关
    if any(keyword in question_lower for keyword in ['成员', '用户', '谁', '人员', '大家', '成员列表']):
        profiles = UserProfile.objects.select_related('user').filter(user__is_superuser=False)
        if profiles:
            data_context.append("【实验室成员列表】")
            for profile in profiles:
                data_context.append(f"- {profile.user.username} ({profile.user.first_name or '未设置姓名'}): {profile.research_topic}")
    
    # 检查是否询问成果相关
    if any(keyword in question_lower for keyword in ['成果', '论文', '专利', '奖项', '发表', 'achievement']):
        achievements = Achievement.objects.all().order_by('-date')[:15]
        if achievements:
            data_context.append("【实验室成果列表】")
            for ach in achievements:
                type_text = {"论文": "📄", "专利": "💡", "奖项": "🏆", "其他": "📌"}.get(ach.achievement_type, "📌")
                data_context.append(f"- {ach.user.username}: {type_text} {ach.title} ({ach.date or '日期未知'})")
    
    # 检查是否询问文档相关
    if any(keyword in question_lower for keyword in ['文档', '文件', '资料', 'document']):
        docs = Document.objects.filter(is_public=True).order_by('-uploaded_at')[:10]
        if docs:
            data_context.append("【公开文档列表】")
            for doc in docs:
                data_context.append(f"- {doc.title} (上传者: {doc.user.username}, {doc.uploaded_at.date()})")
    
    # 检查是否询问公告相关
    if any(keyword in question_lower for keyword in ['公告', '通知', 'announcement']):
        announcements = Announcement.objects.all()[:5]
        if announcements:
            data_context.append("【最新公告】")
            for ann in announcements:
                data_context.append(f"- {ann.title}: {ann.content[:50]}...")
    
    return "\n".join(data_context) if data_context else ""


@login_required
@require_POST
def ai_chat(request):
    """AI 对话接口"""
    try:
        user_message = request.POST.get('message', '')
        if not user_message:
            return JsonResponse({'error': '消息不能为空'}, status=400)
        
        # 根据问题获取相关的实验室数据
        lab_data = get_lab_data_for_ai(user_message)
        
        # 构建系统提示，包含实验室数据
        system_prompt = """你是一个实验室助手，负责回答用户关于学习、科研、工作等方面的问题。回答要简洁明了，友好专业。

实验室数据查询说明：
- 如果用户询问任务完成情况，请根据提供的任务列表统计并分析
- 如果用户询问成员信息，请根据成员列表回答
- 如果用户询问成果，请根据成果列表回答
- 如果用户询问的问题需要数据支持，请先查看下面提供的实验室数据

"""
        if lab_data:
            system_prompt += f"\n以下是实验室的实时数据：\n{lab_data}\n"
        
        # 获取会话历史（从 session 中）
        conversation_history = request.session.get('conversation_history', [])
        
        # 构建请求体
        messages = []
        # 添加系统提示
        messages.append({
            "Role": "system",
            "Content": system_prompt
        })
        # 添加历史对话（最近 6 轮）
        for msg in conversation_history[-6:]:
            messages.append({
                "Role": msg.get("role", "user"),
                "Content": msg.get("content", "")
            })
        # 添加当前问题
        messages.append({
            "Role": "user",
            "Content": user_message
        })
        
        # 调用腾讯云 API
        response = call_tencent_ai(messages)
        
        # 调试：打印响应
        print("API 响应:", response)
        
        # 腾讯云 API 响应在 response['Response'] 中
        api_response = response.get('Response', response)
        
        if api_response and 'Choices' in api_response and len(api_response['Choices']) > 0:
            ai_reply = api_response['Choices'][0]['Message']['Content']
            
            # 更新会话历史
            conversation_history.append({
                "role": "user",
                "content": user_message
            })
            conversation_history.append({
                "role": "assistant",
                "content": ai_reply
            })
            request.session['conversation_history'] = conversation_history
            
            return JsonResponse({
                'success': True,
                'message': ai_reply
            })
        elif 'Error' in api_response:
            error_msg = api_response.get('Error', {}).get('Message', str(api_response))
            return JsonResponse({'error': 'API 错误：' + error_msg}, status=500)
        else:
            return JsonResponse({'error': 'AI 响应异常：' + str(api_response)}, status=500)
            
    except Exception as e:
        print("AI 对话错误:", str(e))
        return JsonResponse({'error': '服务器错误：' + str(e)}, status=500)


def create_excel_report(title, headers, data_rows, filename, statistics=None):
    """创建Excel报告 - 统一格式"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='宋体', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 表头
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    ws.row_dimensions[2].height = 25
    
    # 数据行
    data_font = Font(name='宋体', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, row_data in enumerate(data_rows, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 统计信息
    if statistics:
        stat_row = len(data_rows) + 4
        ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=len(headers))
        stat_cell = ws.cell(row=stat_row, column=1, value=statistics)
        stat_cell.font = Font(name='宋体', size=10, bold=True)
        stat_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col - 1])
        for row in range(3, len(data_rows) + 3):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


def create_plan_excel_with_hierarchy(title, main_plans, filename, show_month=True):
    """创建带有主计划+子计划层级结构的Excel报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel工作表名称最多31个字符
    
    # 表头
    if show_month:
        headers = ['月份', '日期', '计划内容', '子计划', '状态']
    else:
        headers = ['日期', '计划内容', '子计划', '状态']
    
    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='宋体', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 25
    
    # 数据行样式
    data_font = Font(name='宋体', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    row_idx = 3
    total_plans = 0
    completed_plans = 0
    
    for plan in main_plans:
        sub_plans = list(plan.sub_plans.all().order_by('date', 'order'))
        month_str = f"{plan.date.month}月"
        
        # 计算日期范围（优先使用 deadline）
        if plan.deadline:
            date_range = f"{plan.date} 至 {plan.deadline}"
        elif sub_plans:
            end_date = max(sub.date for sub in sub_plans)
            if end_date == plan.date:
                date_range = str(plan.date)
            else:
                date_range = f"{plan.date} 至 {end_date}"
        else:
            date_range = f"{plan.date} 至 至今"
        
        # 子计划内容汇总
        if sub_plans:
            sub_content = '；'.join([sub.content for sub in sub_plans])
        else:
            sub_content = ''
        
        # 状态（使用新的状态判断方法）
        status = plan.get_status_display()
        if plan.is_completed:
            completed_plans += 1
        total_plans += 1
        
        # 主计划行
        col_idx = 1
        if show_month:
            ws.cell(row=row_idx, column=col_idx, value=month_str).font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            col_idx += 1
        
        ws.cell(row=row_idx, column=col_idx, value=date_range).font = data_font
        ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        col_idx += 1
        
        ws.cell(row=row_idx, column=col_idx, value=plan.content).font = data_font
        ws.cell(row=row_idx, column=col_idx).alignment = left_alignment
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        col_idx += 1
        
        ws.cell(row=row_idx, column=col_idx, value=sub_content).font = data_font
        ws.cell(row=row_idx, column=col_idx).alignment = left_alignment
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        col_idx += 1
        
        ws.cell(row=row_idx, column=col_idx, value=status).font = data_font
        ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        
        row_idx += 1
        
        # 子计划行
        for sub in sub_plans:
            col_idx = 1
            if show_month:
                ws.cell(row=row_idx, column=col_idx, value=month_str).font = data_font
                ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
                ws.cell(row=row_idx, column=col_idx).border = thin_border
                col_idx += 1
            
            ws.cell(row=row_idx, column=col_idx, value=str(sub.date)).font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            col_idx += 1
            
            ws.cell(row=row_idx, column=col_idx, value=f"  └ {sub.content}").font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = left_alignment
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            col_idx += 1
            
            ws.cell(row=row_idx, column=col_idx, value='').font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            col_idx += 1
            
            sub_status = sub.get_status_display()
            ws.cell(row=row_idx, column=col_idx, value=sub_status).font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = data_alignment
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            
            row_idx += 1
            total_plans += 1
            if sub.is_completed:
                completed_plans += 1
    
    # 统计信息
    completion_rate = (completed_plans / total_plans * 100) if total_plans > 0 else 0
    stat_row = row_idx + 1
    ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=len(headers))
    stat_cell = ws.cell(row=stat_row, column=1, value=f'统计：总计 {total_plans} 项，已完成 {completed_plans} 项，完成率 {completion_rate:.1f}%')
    stat_cell.font = Font(name='宋体', size=10, bold=True)
    stat_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    ws.column_dimensions['A'].width = 8 if show_month else 12
    ws.column_dimensions['B'].width = 22 if show_month else 35
    ws.column_dimensions['C'].width = 35 if show_month else 40
    ws.column_dimensions['D'].width = 40 if show_month else 10
    if show_month:
        ws.column_dimensions['E'].width = 10
    
    return wb


@login_required
def export_data(request):
    """导出数据"""
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from django.http import HttpResponse
    
    export_type = request.GET.get('type', 'achievements')
    
    if export_type == 'achievements':
        achievements = Achievement.objects.all().select_related('user')
        
        headers = ['标题', '类型', '用户', '日期', '描述']
        data_rows = []
        for ach in achievements:
            data_rows.append([
                ach.title,
                ach.achievement_type,
                ach.user.username,
                str(ach.date) if ach.date else '',
                ach.description or ''
            ])
        
        wb = create_excel_report(
            '成果导出',
            headers,
            data_rows,
            'achievements.xlsx',
            statistics=f'总计: {len(data_rows)} 项'
        )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=achievements.xlsx'
        wb.save(response)
        return response
        
    elif export_type == 'plans':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from calendar import month_name
        
        # 获取所有主计划（parent为None）
        main_plans = DailyPlan.objects.filter(parent__isnull=True).select_related('user').prefetch_related('sub_plans')
        
        wb = Workbook()
        ws = wb.active
        ws.title = '计划导出'
        
        # 标题行
        headers = ['月份', '日期', '计划内容', '子计划', '状态']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value='年报 | 2026年')
        title_cell.font = Font(name='宋体', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[2].height = 25
        
        # 数据行样式
        data_font = Font(name='宋体', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row_idx = 3
        total_plans = 0
        completed_plans = 0
        
        for plan in main_plans:
            sub_plans = list(plan.sub_plans.all().order_by('date', 'order'))
            month_str = f"{plan.date.month}月"
            
            # 计算日期范围
            if sub_plans:
                end_date = max(sub.date for sub in sub_plans)
                if end_date == plan.date:
                    date_range = str(plan.date)
                else:
                    date_range = f"{plan.date} 至 {end_date}"
            else:
                date_range = str(plan.date)
            
            # 子计划内容汇总
            if sub_plans:
                sub_content = '；'.join([sub.content for sub in sub_plans])
            else:
                sub_content = ''
            
            # 状态
            status = '已完成' if plan.is_completed else '进行中'
            if plan.is_completed:
                completed_plans += 1
            total_plans += 1
            
            # 主计划行
            ws.cell(row=row_idx, column=1, value=month_str).font = data_font
            ws.cell(row=row_idx, column=1).alignment = data_alignment
            ws.cell(row=row_idx, column=1).border = thin_border
            
            ws.cell(row=row_idx, column=2, value=date_range).font = data_font
            ws.cell(row=row_idx, column=2).alignment = data_alignment
            ws.cell(row=row_idx, column=2).border = thin_border
            
            ws.cell(row=row_idx, column=3, value=plan.content).font = data_font
            ws.cell(row=row_idx, column=3).alignment = left_alignment
            ws.cell(row=row_idx, column=3).border = thin_border
            
            ws.cell(row=row_idx, column=4, value=sub_content).font = data_font
            ws.cell(row=row_idx, column=4).alignment = left_alignment
            ws.cell(row=row_idx, column=4).border = thin_border
            
            ws.cell(row=row_idx, column=5, value=status).font = data_font
            ws.cell(row=row_idx, column=5).alignment = data_alignment
            ws.cell(row=row_idx, column=5).border = thin_border
            
            row_idx += 1
            
            # 子计划行（如果有）
            for sub in sub_plans:
                ws.cell(row=row_idx, column=1, value=month_str).font = data_font
                ws.cell(row=row_idx, column=1).alignment = data_alignment
                ws.cell(row=row_idx, column=1).border = thin_border
                
                ws.cell(row=row_idx, column=2, value=str(sub.date)).font = data_font
                ws.cell(row=row_idx, column=2).alignment = data_alignment
                ws.cell(row=row_idx, column=2).border = thin_border
                
                ws.cell(row=row_idx, column=3, value=f"  └ {sub.content}").font = data_font
                ws.cell(row=row_idx, column=3).alignment = left_alignment
                ws.cell(row=row_idx, column=3).border = thin_border
                
                ws.cell(row=row_idx, column=4, value='').font = data_font
                ws.cell(row=row_idx, column=4).alignment = data_alignment
                ws.cell(row=row_idx, column=4).border = thin_border
                
                sub_status = '已完成' if sub.is_completed else '进行中'
                ws.cell(row=row_idx, column=5, value=sub_status).font = data_font
                ws.cell(row=row_idx, column=5).alignment = data_alignment
                ws.cell(row=row_idx, column=5).border = thin_border
                
                row_idx += 1
                total_plans += 1
                if sub.is_completed:
                    completed_plans += 1
        
        # 统计信息
        completion_rate = (completed_plans / total_plans * 100) if total_plans > 0 else 0
        stat_row = row_idx + 1
        ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=len(headers))
        stat_cell = ws.cell(row=stat_row, column=1, value=f'年度统计：总计 {total_plans} 项，已完成 {completed_plans} 项，完成率 {completion_rate:.1f}%')
        stat_cell.font = Font(name='宋体', size=10, bold=True)
        stat_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plans_hierarchy.xlsx'
        wb.save(response)
        return response
        
    elif export_type == 'members':
        profiles = UserProfile.objects.select_related('user').filter(user__is_superuser=False)
        
        headers = ['用户名', '姓名', '研究方向', '邮箱', '电话']
        data_rows = []
        for profile in profiles:
            data_rows.append([
                profile.user.username,
                profile.user.first_name or '',
                profile.research_topic or '',
                profile.email or '',
                profile.phone or ''
            ])
        
        wb = create_excel_report(
            '成员导出',
            headers,
            data_rows,
            'members.xlsx',
            statistics=f'总计: {len(data_rows)} 人'
        )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=members.xlsx'
        wb.save(response)
        return response
        
    elif export_type == 'weekly':
        today = timezone.localdate()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        
        # 只获取主计划（parent为None），预加载子计划
        main_plans = DailyPlan.objects.filter(
            user=request.user,
            date__gte=start_of_week,
            date__lte=end_of_week,
            parent__isnull=True
        ).select_related('user').prefetch_related('sub_plans').order_by('date', 'order')
        
        filename = f'周报_{start_of_week}_{end_of_week}.xlsx'
        wb = create_plan_excel_with_hierarchy(
            f'周报 | {start_of_week} 至 {end_of_week}',
            main_plans,
            filename,
            show_month=False
        )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    elif export_type == 'monthly':
        today = timezone.localdate()
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        
        # 只获取主计划（parent为None），预加载子计划
        main_plans = DailyPlan.objects.filter(
            user=request.user,
            date__year=year,
            date__month=month,
            parent__isnull=True
        ).select_related('user').prefetch_related('sub_plans').order_by('date', 'order')
        
        filename = f'月报_{year}年{month}月.xlsx'
        wb = create_plan_excel_with_hierarchy(
            f'月报 | {year}年{month}月',
            main_plans,
            filename,
            show_month=False
        )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    elif export_type == 'yearly':
        today = timezone.localdate()
        year = int(request.GET.get('year', today.year))
        
        # 只获取主计划（parent为None），预加载子计划
        main_plans = DailyPlan.objects.filter(
            user=request.user,
            date__year=year,
            parent__isnull=True
        ).select_related('user').prefetch_related('sub_plans').order_by('date', 'order')
        
        filename = f'年报_{year}年.xlsx'
        wb = create_plan_excel_with_hierarchy(
            f'年报 | {year}年',
            main_plans,
            filename,
            show_month=True
        )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    else:
        return HttpResponse('无效的导出类型')


@login_required
def resume_builder(request):
    """简历生成器"""
    from .forms import ResumeBasicsForm, EducationForm, WorkExperienceForm, SkillForm, ProjectForm

    def collect_resume_data(post):
        resume_data = {
            'resume_type': post.get('resume_type', 'academic'),
            'language': post.get('language', 'zh'),
            'basics': {
                'name': post.get('name', ''),
                'gender': post.get('gender', ''),
                'birth_date': post.get('birth_date', ''),
                'headline': post.get('headline', ''),
                'unit': post.get('unit', ''),
                'email': post.get('email', ''),
                'phone': post.get('phone', ''),
                'location': post.get('location', ''),
                'summary': post.get('summary', ''),
                'photo': post.get('photo_data', ''),
            },
            'education': [],
            'work': [],
            'skills': [],
            'projects': [],
        }

        for i in range(int(post.get('education_count', 0) or 0)):
            prefix = f'education_{i}'
            item = {
                'institution': post.get(f'{prefix}_institution', ''),
                'degree': post.get(f'{prefix}_degree', ''),
                'area': post.get(f'{prefix}_area', ''),
                'start_date': post.get(f'{prefix}_start_date', ''),
                'end_date': post.get(f'{prefix}_end_date', ''),
                'description': post.get(f'{prefix}_description', ''),
            }
            if any(item.values()):
                resume_data['education'].append(item)

        for i in range(int(post.get('work_count', 0) or 0)):
            prefix = f'work_{i}'
            item = {
                'company': post.get(f'{prefix}_company', ''),
                'position': post.get(f'{prefix}_position', ''),
                'start_date': post.get(f'{prefix}_start_date', ''),
                'end_date': post.get(f'{prefix}_end_date', ''),
                'description': post.get(f'{prefix}_description', ''),
            }
            if any(item.values()):
                resume_data['work'].append(item)

        for i in range(int(post.get('skill_count', 0) or 0)):
            prefix = f'skill_{i}'
            item = {
                'category': post.get(f'{prefix}_category', ''),
                'skills': post.get(f'{prefix}_skills', ''),
            }
            if any(item.values()):
                resume_data['skills'].append(item)

        for i in range(int(post.get('project_count', 0) or 0)):
            prefix = f'project_{i}'
            item = {
                'name': post.get(f'{prefix}_name', ''),
                'description': post.get(f'{prefix}_description', ''),
                'link': post.get(f'{prefix}_link', ''),
            }
            if any(item.values()):
                resume_data['projects'].append(item)
        return resume_data

    resume_en_dictionary = {
        '您的姓名': 'Your Name',
        '陈韩宇': 'Chen Hanyu',
        '任千一': 'Ren Qianyi',
        '苏洁': 'Su Jie',
        '张博杭': 'Zhang Bohang',
        '邵荣恒': 'Shao Rongheng',
        '博士研究生': 'Ph.D. Candidate',
        '硕士研究生': 'Master Student',
        '本科生': 'Undergraduate Student',
        '已毕业': 'Alumni',
        '负责人': 'Principal Investigator',
        '男': 'Male',
        '女': 'Female',
        'XX大学 实验室': 'XX University Laboratory',
        'XX大学': 'XX University',
        '实验室': 'Laboratory',
        '北京市 海淀区': 'Haidian District, Beijing',
        '城市，省份': 'City, Province',
        '至今': 'Present',
        '高中': 'High School',
        '专科': 'Associate Degree',
        '本科': "Bachelor's Degree",
        '硕士': "Master's Degree",
        '博士': 'Ph.D.',
        '计算机科学与技术': 'Computer Science and Technology',
        '软件工程': 'Software Engineering',
        '机器学习': 'Machine Learning',
        '深度学习': 'Deep Learning',
        '生物信息学': 'Bioinformatics',
        '图神经网络': 'Graph Neural Networks',
        '表示学习': 'Representation Learning',
        '个人简介': 'Profile',
        '教育经历': 'Education',
        '工作经历': 'Experience',
        '项目经历': 'Projects',
        '科研成果': 'Research Output',
        '技能证书': 'Skills and Certificates',
        '技能': 'Skills',
        '证书': 'Certificates',
        '编程语言': 'Programming Languages',
        '工具框架': 'Tools and Frameworks',
        '工具': 'Tools',
        '框架': 'Frameworks',
        '主修课程': 'Major courses',
        '成就': 'Achievements',
        '工作职责': 'Responsibilities',
        '项目概述': 'Project overview',
        '技术栈': 'Technology stack',
        '您的贡献': 'Contributions',
        '负责': 'Responsible for',
        '参与': 'Participated in',
        '开发': 'Developed',
        '研究': 'Research',
        '系统': 'System',
        '平台': 'Platform',
        '算法': 'Algorithm',
        '数据分析': 'Data Analysis',
        '可视化': 'Visualization',
        '论文': 'Paper',
        '专利': 'Patent',
        '第一作者': 'First Author',
        '通讯作者': 'Corresponding Author',
        '这个人很懒什么也没有留下。': 'No profile summary has been provided.',
    }

    def resume_translate(text, language):
        if not text or language != 'en':
            return text or ''
        translated = str(text)
        for key in sorted(resume_en_dictionary, key=len, reverse=True):
            translated = translated.replace(key, resume_en_dictionary[key])
        return (
            translated
            .replace('，', ', ')
            .replace('。', '. ')
            .replace('；', '; ')
            .replace('：', ': ')
            .replace('、', ', ')
            .replace('（', '(')
            .replace('）', ')')
            .replace('  ', ' ')
            .strip()
        )

    profile = getattr(request.user, 'profile', None)
    default_resume_data = {
        'resume_type': 'academic',
        'language': 'zh',
        'basics': {
            'name': request.user.first_name or request.user.username,
            'gender': '',
            'birth_date': '',
            'headline': profile.get_member_type_display() if profile else '',
            'unit': 'XX大学 实验室',
            'email': (profile.email if profile else '') or request.user.email,
            'phone': profile.phone if profile else '',
            'location': '',
            'summary': profile.bio if profile else '',
            'photo': '',
        },
        'education': [],
        'work': [],
        'skills': [],
        'projects': [],
    }
    
    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'save_draft':
            resume_data = collect_resume_data(request.POST)
            ResumeDraft.objects.update_or_create(
                user=request.user,
                defaults={'data': resume_data}
            )
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'updated_at': timezone.now().strftime('%Y-%m-%d %H:%M')})
            messages.success(request, '简历草稿已保存')
            return redirect('resume_builder')
        
        if action == 'generate_pdf':
            # 收集表单数据
            resume_data = collect_resume_data(request.POST)
            
            # 调试信息
            print(f"收到照片数据：{'有' if resume_data['basics']['photo'] else '无'}")
            if resume_data['basics']['photo']:
                print(f"照片数据长度：{len(resume_data['basics']['photo'])}")
                print(f"照片数据前缀：{resume_data['basics']['photo'][:50]}...")
            
            # 生成 PDF - 与前端预览布局完全一致
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import mm, cm
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from io import BytesIO
                
                # 创建 PDF - 使用较小的边距以确保内容在一页内
                # A4 尺寸: 210mm x 297mm
                buffer = BytesIO()
                margin = 15*mm  # 减小边距以容纳更多内容
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=A4,
                    rightMargin=margin,
                    leftMargin=margin,
                    topMargin=margin,
                    bottomMargin=margin,
                )
                
                # 可用内容宽度 = A4 宽度 (210mm) - 左右边距 (15mm*2) = 180mm
                content_width = 180*mm
                
                elements = []
                styles = getSampleStyleSheet()
                
                # 注册中文字体 - 与预览 CSS 中的 SimSun 一致
                font_name = 'SimSun'
                font_name_bold = 'SimSun'
                font_registered = False
                font_paths = [
                    'C:/Windows/Fonts/simsun.ttc',
                    'C:/Windows/Fonts/SimSun.ttf',
                    '/usr/share/fonts/truetype/SimSun.ttf',
                ]
                for fp in font_paths:
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, fp))
                        font_registered = True
                        break
                    except:
                        continue
                
                if not font_registered:
                    font_name = 'Helvetica'
                    font_name_bold = 'Helvetica-Bold'
                
                # ===== 样式定义 - 与预览 CSS 像素级对应 =====
                
                # 姓名 - 对应 .resume-name: font-size:26px, font-weight:700, color:#0f172a
                title_style = ParagraphStyle(
                    'ResumeTitle',
                    parent=styles['Normal'],
                    fontSize=26,
                    textColor=colors.HexColor('#0f172a'),
                    fontName=font_name_bold if not font_registered else font_name,
                    alignment=1,
                    leading=34,
                    spaceAfter=4,
                )
                
                # 职位 - 对应 .resume-headline: font-size:14px, color:#64748b
                headline_style = ParagraphStyle(
                    'ResumeHeadline',
                    parent=styles['Normal'],
                    fontSize=14,
                    textColor=colors.HexColor('#64748b'),
                    fontName=font_name,
                    alignment=1,
                    leading=18,
                    spaceAfter=4,
                )
                
                # 联系信息 - 对应 .resume-contact: font-size:10px, color:#64748b
                contact_style = ParagraphStyle(
                    'ResumeContact',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#64748b'),
                    fontName=font_name,
                    alignment=1,
                    leading=14,
                    spaceBefore=8,
                    spaceAfter=0,
                )
                
                # 章节标题 - 对应 .resume-section-title: font-size:16px, font-weight:700, color:#0f172a
                section_title_style = ParagraphStyle(
                    'ResumeSectionTitle',
                    parent=styles['Normal'],
                    fontSize=16,
                    textColor=colors.HexColor('#0f172a'),
                    fontName=font_name_bold if not font_registered else font_name,
                    leading=20,
                    spaceBefore=10,  # 减小章节间距
                    spaceAfter=2,
                )
                
                # 条目标题 - 对应 .resume-entry-title: font-size:11px, font-weight:700, color:#0f172a
                entry_title_style = ParagraphStyle(
                    'ResumeEntryTitle',
                    parent=styles['Normal'],
                    fontSize=11,
                    textColor=colors.HexColor('#0f172a'),
                    fontName=font_name_bold if not font_registered else font_name,
                    leading=14,
                )
                
                # 条目日期 - 对应 .resume-entry-date: font-size:10px, color:#64748b, italic
                entry_date_style = ParagraphStyle(
                    'ResumeEntryDate',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#64748b'),
                    fontName=font_name,
                    leading=14,
                    alignment=2,  # 右对齐
                )
                
                # 条目副标题 - 对应 .resume-entry-subtitle: font-size:10px, color:#64748b, italic
                entry_subtitle_style = ParagraphStyle(
                    'ResumeEntrySubtitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#64748b'),
                    fontName=font_name,
                    leading=14,
                    spaceAfter=4,
                )
                
                # 正文/描述 - 对应 .resume-entry-description / .resume-summary-text: font-size:10px, color:#475569
                body_style = ParagraphStyle(
                    'ResumeBody',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#475569'),
                    fontName=font_name,
                    leading=14,  # 减小行高
                )
                
                # ===== 辅助函数 =====
                def add_section_divider():
                    """添加章节分隔线 - 对应 CSS border-bottom: 2px solid #e2e8f0"""
                    elements.append(HRFlowable(
                        width="100%",
                        thickness=2,
                        color=colors.HexColor('#e2e8f0'),
                        spaceBefore=0,
                        spaceAfter=6,
                    ))
                
                def add_entry_header(title_text, date_text=''):
                    """添加条目头部 - 左侧标题+右侧日期，对应预览中的 flex 布局"""
                    if date_text:
                        header_table = Table(
                            [[Paragraph(f'<b>{title_text}</b>', entry_title_style),
                              Paragraph(f'<i>{date_text}</i>', entry_date_style)]],
                            colWidths=[content_width * 0.7, content_width * 0.3],
                        )
                        header_table.setStyle(TableStyle([
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('TOPPADDING', (0, 0), (-1, -1), 0),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                            ('LEFTPADDING', (0, 0), (-1, -1), 0),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                        ]))
                        elements.append(header_table)
                    else:
                        elements.append(Paragraph(f'<b>{title_text}</b>', entry_title_style))
                
                # ===== 构建 PDF 内容 =====
                basics = resume_data['basics']
                is_english_resume = resume_data.get('language') == 'en'
                section_labels = {
                    'summary': 'Profile' if is_english_resume else '个人简介',
                    'education': 'Education' if is_english_resume else '教育经历',
                    'work': 'Experience' if is_english_resume else '工作经历',
                    'skills': 'Skills' if is_english_resume else '技能',
                    'projects': 'Projects' if is_english_resume else '项目经历',
                    'present': 'Present' if is_english_resume else '至今',
                    'filename_suffix': 'Resume' if is_english_resume else '简历',
                }
                tr = lambda value: resume_translate(value, resume_data.get('language', 'zh'))
                
                # 先构建联系信息列表（无论是否有照片都需要）
                contact_parts = []
                if basics.get('gender'):
                    contact_parts.append(tr(basics['gender']))
                if basics.get('unit'):
                    contact_parts.append(tr(basics['unit']))
                if basics['email']:
                    contact_parts.append(basics['email'])
                if basics['phone']:
                    contact_parts.append(basics['phone'])
                if basics['location']:
                    contact_parts.append(tr(basics['location']))
                if basics.get('birth_date'):
                    contact_parts.append(basics['birth_date'])
                contact_text = " | ".join(contact_parts) if contact_parts else ""
                
                # 处理照片（如果有）
                photo_img_element = None
                if basics.get('photo'):
                    try:
                        from reportlab.platypus import Image
                        from reportlab.lib.utils import ImageReader
                        import base64
                        from io import BytesIO
                        # 解析 base64 照片数据
                        photo_data = basics['photo']
                        if photo_data.startswith('data:image'):
                            # 移除 data:image/xxx;base64, 前缀
                            img_data = photo_data.split(',')[1]
                            img_bytes = base64.b64decode(img_data)
                            # 使用 BytesIO 创建文件-like 对象
                            img_buffer = BytesIO(img_bytes)
                            photo_img_element = Image(img_buffer)
                            print(f"照片处理成功，大小: {len(img_bytes)} bytes")
                    except Exception as e:
                        print(f"处理照片失败：{e}")
                        import traceback
                        print(traceback.format_exc())
                
                # 基本信息区域（带照片）
                if photo_img_element:
                    # 使用表格布局：左侧信息，右侧照片
                    from reportlab.platypus import Table
                    photo_width = 25*mm  # 80px ≈ 25mm
                    photo_height = 31*mm  # 100px ≈ 31mm
                    
                    # 设置图片尺寸
                    photo_img_element.drawWidth = photo_width
                    photo_img_element.drawHeight = photo_height
                    
                    # 创建左侧信息文本（将多个段落合并为一个）
                    left_text = ""
                    if basics['name']:
                        left_text += f"<font size='26' color='#0f172a'><b>{tr(basics['name'])}</b></font><br/>"
                    if basics['headline']:
                        left_text += f"<font size='14' color='#64748b'>{tr(basics['headline'])}</font><br/>"
                    if contact_text:
                        left_text += f"<font size='10' color='#64748b'>{contact_text}</font>"
                    
                    # 构建表格数据 - 左侧是段落，右侧是图片
                    left_para = Paragraph(left_text, ParagraphStyle(
                        'LeftInfo',
                        parent=styles['Normal'],
                        fontName=font_name,
                        leading=18,
                        alignment=0,  # 左对齐
                    ))
                    
                    table_data = [[left_para, photo_img_element]]
                    
                    info_table = Table(
                        table_data,
                        colWidths=[content_width - photo_width - 5*mm, photo_width],
                        rowHeights=[photo_height]
                    )
                    info_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                        ('LEFTPADDING', (0, 0), (0, 0), 0),
                        ('RIGHTPADDING', (0, 0), (0, 0), 5*mm),
                        ('LEFTPADDING', (1, 0), (1, 0), 5*mm),
                        ('RIGHTPADDING', (1, 0), (1, 0), 0),
                    ]))
                    elements.append(info_table)
                else:
                    # 没有照片，正常布局
                    # 姓名
                    if basics['name']:
                        elements.append(Paragraph(tr(basics['name']), title_style))
                    
                    # 职位/头衔
                    if basics['headline']:
                        elements.append(Paragraph(tr(basics['headline']), headline_style))
                    
                    # 联系信息 - 与预览一致使用 | 分隔
                    if contact_parts:
                        elements.append(Paragraph(" | ".join(contact_parts), contact_style))
                
                # 头部分隔线 - 对应 .resume-header 的 border-bottom
                elements.append(Spacer(1, 8))
                elements.append(HRFlowable(
                    width="100%",
                    thickness=2,
                    color=colors.HexColor('#e2e8f0'),
                    spaceBefore=0,
                    spaceAfter=16,
                ))
                
                # 个人简介 - 对应预览中的"个人简介"章节
                if basics['summary']:
                    elements.append(Paragraph(section_labels['summary'], section_title_style))
                    add_section_divider()
                    elements.append(Paragraph(tr(basics['summary']), body_style))
                
                # 教育经历
                if resume_data['education']:
                    elements.append(Paragraph(section_labels['education'], section_title_style))
                    add_section_divider()
                    
                    for edu in resume_data['education']:
                        # 日期字符串
                        date_str = ''
                        if edu['start_date'] and edu['end_date']:
                            date_str = f"{edu['start_date']} - {edu['end_date']}"
                        elif edu['start_date']:
                            date_str = f"{edu['start_date']} - {section_labels['present']}"
                        
                        add_entry_header(tr(edu['institution']), date_str)
                        
                        # 副标题: 学位 - 专业
                        subtitle = f"{tr(edu['degree'])} - {tr(edu['area'])}"
                        elements.append(Paragraph(f'<i>{subtitle}</i>', entry_subtitle_style))
                        
                        if edu['description']:
                            elements.append(Paragraph(tr(edu['description']), body_style))
                        
                        elements.append(Spacer(1, 6))
                
                # 工作经历
                if resume_data['work']:
                    elements.append(Paragraph(section_labels['work'], section_title_style))
                    add_section_divider()
                    
                    for work in resume_data['work']:
                        date_str = ''
                        if work['start_date'] and work['end_date']:
                            date_str = f"{work['start_date']} - {work['end_date']}"
                        elif work['start_date']:
                            date_str = f"{work['start_date']} - {section_labels['present']}"
                        
                        add_entry_header(tr(work['company']), date_str)
                        elements.append(Paragraph(f"<i>{tr(work['position'])}</i>", entry_subtitle_style))
                        
                        if work['description']:
                            elements.append(Paragraph(tr(work['description']), body_style))
                        
                        elements.append(Spacer(1, 6))
                
                # 技能
                if resume_data['skills']:
                    elements.append(Paragraph(section_labels['skills'], section_title_style))
                    add_section_divider()
                    
                    for skill in resume_data['skills']:
                        skill_text = f"<b>{tr(skill['category'])}:</b> {tr(skill['skills'])}"
                        elements.append(Paragraph(skill_text, body_style))
                        elements.append(Spacer(1, 3))
                
                # 项目经历
                if resume_data['projects']:
                    elements.append(Paragraph(section_labels['projects'], section_title_style))
                    add_section_divider()
                    
                    for project in resume_data['projects']:
                        project_title = tr(project['name'])
                        if project['link']:
                            project_title += f" ({project['link']})"
                        
                        elements.append(Paragraph(f'<b>{project_title}</b>', entry_title_style))
                        
                        if project['description']:
                            elements.append(Paragraph(tr(project['description']), body_style))
                        
                        elements.append(Spacer(1, 6))
                
                # 构建 PDF
                doc.build(elements)
                buffer.seek(0)
                
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                filename_name = tr(basics['name']).replace(' ', '_') if basics['name'] else 'resume'
                filename_name = ''.join(ch for ch in filename_name if ch.isascii() and (ch.isalnum() or ch in {'_', '-'})).strip('_-')
                if not filename_name:
                    filename_name = 'resume'
                filename = f"{filename_name}_{section_labels['filename_suffix']}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                # 禁用缓存以确保每次下载都是最新的
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                return response
                
            except Exception as e:
                import traceback
                error_msg = f'生成 PDF 失败：{str(e)}'
                print(f"PDF 生成错误：{traceback.format_exc()}")
                
                # 检查是否是 AJAX 请求
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': error_msg}, status=500)
                else:
                    messages.error(request, error_msg)
                    return redirect('resume_builder')
    
    draft = ResumeDraft.objects.filter(user=request.user).first()
    resume_initial = draft.data if draft and draft.data else default_resume_data

    # GET 请求，显示表单
    context = {
        'basics_form': ResumeBasicsForm(),
        'education_form': EducationForm(),
        'work_form': WorkExperienceForm(),
        'skill_form': SkillForm(),
        'project_form': ProjectForm(),
        'resume_initial': resume_initial,
        'resume_initial_json': json.dumps(resume_initial, ensure_ascii=False),
        'resume_draft': draft,
    }
    return render(request, 'lab_management/resume_builder.html', context)


def call_tencent_ai(messages):
    """调用腾讯云混元大模型 API"""
    import time
    import hashlib
    import hmac
    from datetime import datetime
    import json
    
    # 腾讯云 API 配置
    secret_id = settings.TENCENT_AI_SECRET_ID
    secret_key = settings.TENCENT_AI_SECRET_KEY
    service = "hunyuan"
    host = "hunyuan.tencentcloudapi.com"
    action = "ChatCompletions"
    version = "2023-09-01"
    region = "ap-guangzhou"
    model = settings.TENCENT_AI_MODEL
    
    # 构建请求体
    payload = {
        "Model": model,
        "Messages": messages
    }
    
    # 准备签名
    timestamp = int(time.time())
    date = datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%d")
    
    # 构建规范请求
    http_request_method = "POST"
    canonical_uri = "/"
    canonical_querystring = ""
    ct = "application/json; charset=utf-8"
    canonical_headers = "content-type:%s\nhost:%s\n" % (ct, host)
    signed_headers = "content-type;host"
    hashed_request_payload = hashlib.sha256(json.dumps(payload, ensure_ascii=False).encode('utf-8')).hexdigest()
    canonical_request = (http_request_method + "\n" +
                         canonical_uri + "\n" +
                         canonical_querystring + "\n" +
                         canonical_headers + "\n" +
                         signed_headers + "\n" +
                         hashed_request_payload)
    
    # 构建待签名字符串
    algorithm = "TC3-HMAC-SHA256"
    credential_scope = date + "/" + service + "/" + "tc3_request"
    hashed_canonical_request = hashlib.sha256(canonical_request.encode('utf-8')).hexdigest()
    string_to_sign = (algorithm + "\n" +
                      str(timestamp) + "\n" +
                      credential_scope + "\n" +
                      hashed_canonical_request)
    
    # 计算签名
    def sign(key, msg):
        return hmac.new(key, msg.encode('utf-8'), hashlib.sha256).digest()
    
    secret_date = sign(("TC3" + secret_key).encode('utf-8'), date)
    secret_service = sign(secret_date, service)
    secret_signing = sign(secret_service, "tc3_request")
    signature = hmac.new(secret_signing, string_to_sign.encode('utf-8'), hashlib.sha256).hexdigest()
    
    # 构建授权头
    authorization = (algorithm + " " +
                     "Credential=" + secret_id + "/" + credential_scope + ", " +
                     "SignedHeaders=" + signed_headers + ", " +
                     "Signature=" + signature)
    
    # 发送请求
    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Host": host,
        "X-TC-Action": action,
        "X-TC-Version": version,
        "X-TC-Region": region,
        "X-TC-Timestamp": str(timestamp),
        "Authorization": authorization
    }
    
    print("正在调用腾讯云 API...")
    print("请求头:", {k: v[:50] + "..." if len(str(v)) > 50 else v for k, v in headers.items()})
    print("请求体:", json.dumps(payload, ensure_ascii=False)[:200])
    
    try:
        response = requests.post(
            "https://" + host,
            headers=headers,
            data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),
            timeout=30
        )
        print("HTTP 状态码:", response.status_code)
        print("响应内容:", response.text[:500])
        return response.json()
    except Exception as e:
        print("API 调用异常:", str(e))
        raise Exception(f"API 调用失败：{str(e)}")

